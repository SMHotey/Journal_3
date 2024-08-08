from django.shortcuts import render, redirect, get_object_or_404
from .forms import UploadFileForm
from openpyxl import load_workbook
import re

from .models import Order


def delete_order(request, order_id):
    order = get_object_or_404(Order, number=order_id)
    if request.method == 'POST':
        order.delete()
        return redirect('orders_list')  # переадресация на список заказов
    return render(request, 'orders/delete_order.html', {'order': order})


def order_view(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES['file']

            wb = load_workbook(uploaded_file)
            sheet = wb.active

            doors_nk, doors_2nk, hatches_nk = 0, 0, 0
            doors_sk, doors_2sk, hatches_sk, gates = 0, 0, 0, 0
            glass_order = {}

            max_row, cur_column = 9, 15
            line = []
            order = []

            class Position:
                def __init__(self, num_position, name, high, length, active_length, direction, trim,
                             accessories, closer, doorstep, ral, quantity, comment, *glasses):
                    self.num_position = num_position
                    self.name = name
                    self.high = high
                    self.length = length
                    self.active_length = active_length
                    self.direction = direction
                    self.trim = trim
                    self.accessories = accessories
                    self.closer = closer
                    self.doorstep = doorstep
                    self.ral = ral
                    self.quantity = quantity
                    self.comment = comment
                    self.glasses = list(glasses)

                def __str__(self):
                    return (
                        f'{self.num_position}. {self.name}, размер по коробке ({self.high}, {self.length}), {self.direction},'
                        f' наличник: {self.trim}, фурнитура: {self.accessories}, {self.glasses}')

            def get_value(row, column):
                return sheet.cell(row=row, column=column).value

            # Определяем размер таблицы с данными
            while get_value(max_row, cur_column) != 'шт.':
                max_row += 1

            # Читаем файл с заявкой
            seq = [1, 2, 3, 4, 5, 6, 9, 10, 11, 12, 13, 14, 15, 7, 8]
            for row in range(8, max_row):
                line.clear()
                for column in seq:
                    line.append(get_value(row, column))

                if get_value(row, 1):
                    order.append(Position(*line))
                else:
                    add_glass = (get_value(row, 7), get_value(row, 8))
                    order[-1].glasses += add_glass

            for i in range(0, len(order)):
                quantity = order[i].quantity
                name = order[i].name
                if re.search('дверь', name.lower()):
                    if re.search('-м', name.lower()):
                        if order[i].active_length:
                            doors_2nk += quantity
                        else:
                            doors_nk += quantity
                    elif order[i].active_length:
                        doors_2sk += quantity
                    else:
                        doors_sk += quantity

                if re.search('ворота', name.lower()):
                    gates += order[i].quantity
                if re.search('люк', name.lower()):
                    if re.search('-м', name.lower()):
                        hatches_nk += quantity
                    else:
                        hatches_sk += quantity

                if order[i].glasses[0]:
                    for glass in range(0, len(order[i].glasses), 2):
                        key = (order[i].glasses[glass], order[i].glasses[glass + 1])
                        glass_order[str(key)] = glass_order.get(str(key), 0) + order[i].quantity

            order = Order(
                doors_nk=doors_nk,
                doors_sk=doors_sk,
                doors_2nk=doors_2nk,
                doors_2sk=doors_2sk,
                hatches_nk=hatches_nk,
                hatches_sk=hatches_sk,
                gates=gates,
                glass_order=str(glass_order)
            )
            order.save()
            return render(request, 'success.html')

    form = UploadFileForm()
    return render(request, 'upload.html', {'form': form})


def orders_list(request):
    orders = Order.objects.all()
    return render(request, 'orders_list.html', {'orders': orders})
